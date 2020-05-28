# -*-  coding: utf-8 -*-
# Author: cao wang
# Datetime : 2020
# software: PyCharm
# 收获:
from 项目.项目二_文书内容提取.提取模板模块 import jiebacut_word as j
import openpyxl

#pprint.pprint(j.jieba_cutword())

path = 'I:\迅雷下载\捕诉\_捕诉合一_的域外实践及其启示_孙长永.pdf'
words = j.jieba_cutword(path)


def openxl():
    """写入xlsx文档"""
    wb = openpyxl.Workbook()
    ws = wb.active

    for i in range(len(words)):
        ws.cell(row=1, column=1, value='索引')
        ws.cell(row=1, column=2, value='分词')
        ws.cell(row=1, column=3, value='频率')
        ws.cell(row=1, column=4, value='本文作者')
        ws.cell(row=i + 2, column=1, value=i)
        ws.cell(row=i+2, column=2, value=words[i][0])
        ws.cell(row=i+2, column=3, value=words[i][1])
        ws.cell(row=i + 2, column=4, value='孙常用')


    wb.save('test5.xlsx')  # 这个表不可以打开状态下添加
openxl()
import plotly.express as px
def openxl_df():
    """写入xlsx文档"""
    for df in px.data.gapminder():

        with open('test.txt','a')as f:
            f.write(str(df))
            f.write(str(df))
            f.write(str(df))
            f.write(str(df))
            f.write(str(df))

#openxl_df()


