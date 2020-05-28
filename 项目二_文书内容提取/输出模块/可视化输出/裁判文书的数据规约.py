# -*-  coding: utf-8 -*-
# Author: cao wang
# Datetime : 2020
# software: PyCharm
# 收获:
import openpyxl
from 项目.项目二_文书内容提取.输入模块.数据库api接口.mongodb数据库接口 import database_connect as db
import pandas as pd
import glob
import os


class Caipanwenshu():
    """mongodb数据库的数据规约化并写入xlsx中"""

    def __init__(self,port,input_file,output_file):
        #self.data = data
        self.port = port
        '''端口不是必须也没有什么用只是为了连接,控制函数的调用'''

        self.input_file = input_file
        self.output_file = output_file
        """输入输出文件"""



    def mongodb_data_clean(self):
        """mongodb的数据先清洗为可以写入xlsx的格式"""
        try:
            index = 0
            for i in db(self.port):
                # print('数据名称:{name}'.format(name=i))
                print()
                index += 1

                print('        这是第%d个整理后的数据！！！！！！！！！！！！！！！！！！！！！！！！' % index)

                '''，前为yield,列表内的元组格式内容列表化'''
                for title, data in i.items():
                    data_list = [list(datas) for datas in data]
                    yield [title, data_list]
        except:
            pass

        #print('         一共%d个数据----------------------------------------------------' % index)
        #直接print不出是由于新增title是一个字符串，不能遍历


    def openxl(self):
        """写入xlsx文档"""
        wb = openpyxl.Workbook()
        ws = wb.active
        index = 0
        for datas in self.mongodb_data_clean():
            index += 1

            for i in range(len(datas[1])):

                ws.cell(row=1, column=1, value='分词')
                ws.cell(row=1, column=2, value='频率')
                ws.cell(row=1, column=3, value='文书名')

                ws.cell(row=i + 1, column=1, value=datas[1][i][0])
                ws.cell(row=i + 1, column=2, value=datas[1][i][1])
                ws.cell(row=i + 1, column=3, value=datas[0])

            wb.save('J:\PyCharm项目\项目\项目二_文书内容提取\输出模块\可视化输出\\xslx写入\\test%d.xlsx' % index)  # 这个表不可以打开状态下添加
        print('xlsx写入成功--------------------------------------------------------')


    def xlsx_concat(self):
        """xlsx文档合并"""

        try:
            input_files = 'J:\PyCharm项目\项目\项目二_文书内容提取\输出模块\可视化输出\\xslx写入'
            print('读取到了第一个文件.............................................')
            output_files = self.input_file
            # 需要创建，否则没有输出对象
            print('读取到了第二个文件.............................................')
            all_workbook = glob.glob(os.path.join(input_files, '*.xlsx'))
            # print(all_workbook)

            data_frames = []
            '''重点来了，遍历工作簿、工作表，簿为列表，表为字典'''
            for workbook in all_workbook:
                all_sheets = pd.read_excel(workbook, sheet_name=None, index_col=None)
                for worksheet_name, data in all_sheets.items():
                    data_frames.append(data)  # sheets内容为表名与表内容

            all_data = pd.concat(data_frames)  # axis即竖连接还是横连接
            # print(data_frames)
            writer = pd.ExcelWriter(output_files)
            all_data[:1048576].to_excel(writer, sheet_name='all_output', index=False)
            all_data[1048576:].to_excel(writer, sheet_name='new_output', index=False)
            writer.save()
        except Exception as e:
            print(e)






    def drop_duplicate(self):
        """好像不能直接去重，需要专门构建一个函数"""

        file = self.input_file
        data_frame = pd.read_excel(file,index_col=None)
        print(type(data_frame))
        data_frames = data_frame.drop_duplicates(keep=False)
        print(data_frames)
        writer = pd.ExcelWriter(self.output_file)
        data_frames.to_excel(writer, sheet_name='all_output', index=False)
        writer.save()

input_file = "J:\PyCharm项目\项目\项目二_文书内容提取\输出模块\可视化输出\\规约完成文件\\test汇总表.xlsx"
output_file = "J:\PyCharm项目\项目\项目二_文书内容提取\输出模块\可视化输出\\规约完成文件\\test汇总表(去重）.xlsx"
Caipanwenshu(27017,input_file,output_file).drop_duplicate()
